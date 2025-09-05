# Copyright (c) 2025 The OSCAL Compass Authors. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Transform HIPAA cprt_SP800_66_2_0_0_08-02-2025.xlsx to OSCAL catalog."""

import openpyxl
import json
import uuid
from datetime import datetime
import argparse

def clean_prop_value(value):
    if not value:
        return None
    value = value.strip()
    value = '\n'.join([line.strip() for line in value.split('\n') if line.strip()])
    value = value.replace('\n', ', ')
    return value

def create_oscal_catalog(input_file, output_file, title, version, oscal_version):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    columns = [cell.value for cell in ws[1]]

    catalog = {
        "catalog": {
            "uuid": str(uuid.uuid4()),
            "metadata": {
                "title": title,
                "version": version,
                "oscal-version": oscal_version,
                "last-modified": datetime.now().isoformat()
            },
            "groups": []
        }
    }

    level1_groups = {}
    level2_groups = {}
    control_counter = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        level1_title = row[1]
        if level1_title:
            level1_title, level1_desc = level1_title.split(":", 1)
            level1_id = f"group-{len(level1_groups) + 1}"
            level1_group = {
                "id": level1_id,
                "title": level1_title.strip(),
                "props": [],
                "parts": [
                    {
                        "id": f"{level1_id}_desc",
                        "name": "description",
                        "prose": level1_desc.strip()
                    }
                ],
                "groups": []
            }
            if row[0]:
                prop_name = columns[0].lower().replace(" ", "-")
                prop_value = clean_prop_value(row[0])
                if prop_value:
                    level1_group["props"].append({
                        "name": prop_name,
                        "value": prop_value,
                        "remarks": columns[0]
                    })
            level1_groups[level1_id] = level1_group
            catalog["catalog"]["groups"].append(level1_group)

        level2_title = row[3]
        if level2_title:
            level2_title, level2_desc = level2_title.split(":", 1)
            level2_id = f"{level1_id}.{len(level2_groups) + 1}"
            level2_group = {
                "id": level2_id,
                "title": level2_title.strip(),
                "props": [],
                "parts": [
                    {
                        "id": f"{level2_id}_desc",
                        "name": "description",
                        "prose": level2_desc.strip()
                    }
                ],
                "controls": []
            }
            if row[2]:
                prop_name = columns[2].lower().replace(" ", "-")
                prop_value = clean_prop_value(row[2])
                if prop_value:
                    level2_group["props"].append({
                        "name": prop_name,
                        "value": prop_value,
                        "remarks": columns[2]
                    })
            level2_groups[level2_id] = level2_group
            level1_group["groups"].append(level2_group)

        control_title = row[4]
        if control_title and control_title != "Implementation Specification (Required)":
            control_id = f"hipaa-{control_counter:03}"
            control = {
                "id": control_id,
                "title": control_title.strip(),
                "parts": []
            }
            if row[5]:
                control["parts"].append({
                    "id": f"{control_id}_smt",
                    "name": "statement",
                    "prose": row[5].strip()
                })
            if row[6]:
                control["parts"].append({
                    "id": f"{control_id}_qst",
                    "name": "sample_questions",
                    "prose": row[6].strip()
                })
            level2_group["controls"].append(control)
            control_counter += 1

    with open(output_file, 'w') as f:
        json.dump(catalog, f, indent=4)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="The path to the consumed xlsx file")
    parser.add_argument("--output", required=True, help="The path to the produced OSCAL catalog file")
    parser.add_argument("--title", required=True, help="The title of the produced OSCAL catalog")
    parser.add_argument("--version", required=True, help="The version of the produced OSCAL catalog")
    parser.add_argument("--oscal-version", required=True, help="The oscal-version of the produced OSCAL catalog")
    args = parser.parse_args()

    create_oscal_catalog(args.input, args.output, args.title, args.version, args.oscal_version)